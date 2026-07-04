"""Test de exportacion (P2): CSV tidy, XLSX multi-hoja y PDF."""
from __future__ import annotations

from app.services import curso_stats_service as css
from app.services import export_service as ex

PAUTA = {1: "A", 2: "B", 3: "C", 4: "D"}
ALUMNOS = [
    {"student_id": "S1", "respuestas": {1: "A", 2: "B", 3: "C", 4: "D"}},
    {"student_id": "S2", "respuestas": {1: "A", 2: "B", 3: "C", 4: "A"}},
    {"student_id": "S3", "respuestas": {1: "A", 2: "A", 3: "A", 4: "A"}},
]
TE = {i: {"ra": f"RA{i}", "bloom": "Comprension", "unidad": "Unidad 1"} for i in (1, 2, 3, 4)}


def _resultado():
    return css.analizar_evaluacion(ALUMNOS, PAUTA, te_tags=TE)


def test_csv_tidy():
    csv_txt = ex.dataset_csv(_resultado())
    lineas = csv_txt.strip().splitlines()
    assert lineas[0].startswith("student_id,item,ra")
    assert len(lineas) == 1 + 3 * 4  # header + alumno x item


def test_xlsx_multi_hoja(tmp_path):
    import openpyxl
    p = tmp_path / "curso.xlsx"
    ex.to_xlsx(_resultado(), str(p))
    wb = openpyxl.load_workbook(str(p))
    assert wb.sheetnames == ["Resumen", "Items", "Nomina", "Datos_largo"]
    assert wb["Datos_largo"].max_row == 1 + 3 * 4
    assert wb["Nomina"].max_row == 1 + 3


def test_pdf_generado(tmp_path):
    p = tmp_path / "curso.pdf"
    ex.to_pdf(_resultado(), str(p), titulo="Test")
    data = p.read_bytes()
    assert data[:4] == b"%PDF"
    assert len(data) > 1000
