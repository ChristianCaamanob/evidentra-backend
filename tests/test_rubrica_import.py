"""
Test del import de rubrica desde .xlsx: parser (estructura analitica, filtra basura) y el
endpoint (preview sin escribir + confirmar crea los criterios).
"""
from __future__ import annotations

import io

import app.models.course  # noqa: F401
import app.models.teacher  # noqa: F401
import app.models.student  # noqa: F401
import app.models.assessment  # noqa: F401
import app.models.answer_key  # noqa: F401
import app.models.scan  # noqa: F401
import app.models.result  # noqa: F401
import app.models.feedback  # noqa: F401
import app.models.password_reset  # noqa: F401
import app.models.curriculo  # noqa: F401
import app.models.validacion  # noqa: F401
import app.models.aprendizaje  # noqa: F401

import openpyxl
import pytest
from fastapi.testclient import TestClient
from sqlalchemy import create_engine
from sqlalchemy.orm import Session, sessionmaker
from sqlalchemy.pool import StaticPool

from app.main import app
from app.api.deps import get_db, usuario_actual
from app.models.base import Base
from app.models.course import Course
from app.models.assessment import Assessment
from app.models.answer_key import AnswerKey, AnswerKeyItem, RubricCriterion
from app.services.rubrica_import_service import parse_rubrica_xlsx

_CREADOR = type("U", (), {"rol": "creador"})()


def _xlsx_rubrica():
    wb = openpyxl.Workbook(); ws = wb.active
    ws["C1"] = "Contenido de la presentación (Puntaje Grupal)"     # seccion sobre el encabezado
    ws["C2"] = "Categoría"
    ws["D2"] = "Excelente (3 puntos)"; ws["E2"] = "Bueno (2 puntos)"
    ws["F2"] = "Regular (1 punto)"; ws["G2"] = "Deficiente (0 punto)"
    ws["H2"] = "Factor de multiplicación"; ws["I2"] = "Puntaje"
    ws["C3"] = "Integración"; ws["D3"] = "claro"; ws["E3"] = "adecuado"
    ws["F3"] = "parcial"; ws["G3"] = "incompleto"; ws["H3"] = 2
    ws["C4"] = "Calidad"; ws["D4"] = "todo"; ws["E4"] = "80%"; ws["F4"] = "60%"; ws["G4"] = "menos"; ws["H4"] = 1
    ws["C5"] = "Preguntas finales (Puntaje Individual)"            # seccion individual
    ws["C6"] = "Respuesta"; ws["D6"] = "correcta"; ws["E6"] = "ok"; ws["F6"] = "parcial"; ws["G6"] = "mala"; ws["H6"] = 1
    ws["C7"] = 57; ws["E7"] = 57; ws["F7"] = 7                     # fila espuria (tabla conversion)
    b = io.BytesIO(); wb.save(b); return b.getvalue()


def test_parser_extrae_criterios_y_filtra_basura():
    p = parse_rubrica_xlsx(_xlsx_rubrica())
    assert p["n_criterios"] == 3                       # Integracion, Calidad, Respuesta (no el "57")
    assert p["escala"] == [{"nivel": "Excelente", "puntos": 3}, {"nivel": "Bueno", "puntos": 2},
                           {"nivel": "Regular", "puntos": 1}, {"nivel": "Deficiente", "puntos": 0}]
    integ = p["criterios"][0]
    assert integ["name"] == "Integración" and integ["weight"] == 2.0
    assert integ["ambito"] == "grupal" and "Contenido" in integ["seccion"]
    assert [n["puntos"] for n in integ["niveles"]] == [3, 2, 1, 0]
    assert integ["niveles"][0]["descriptor"] == "claro"
    # el ultimo criterio es individual
    assert p["criterios"][-1]["ambito"] == "individual"


def test_parser_rechaza_archivo_no_rubrica():
    wb = openpyxl.Workbook(); ws = wb.active
    ws["A1"] = "hola"; ws["A2"] = "mundo"
    b = io.BytesIO(); wb.save(b)
    with pytest.raises(Exception):
        parse_rubrica_xlsx(b.getvalue())


@pytest.fixture()
def entorno():
    engine = create_engine("sqlite://", connect_args={"check_same_thread": False},
                           poolclass=StaticPool)
    Base.metadata.create_all(engine)
    with Session(engine) as s:
        course = Course(name="Morfologia", code="DMOR0030",
                        grading_scale="chile_1_7", passing_threshold=60.0)
        s.add(course); s.commit(); s.refresh(course)
        a = Assessment(course_id=course.id, name="Infografia", modalidad="oral",
                       grading_scale="chile_1_7", passing_threshold=60.0)
        s.add(a); s.commit(); s.refresh(a)
        ak = AnswerKey(assessment_id=a.id, status="valid", is_valid=True)
        s.add(ak); s.commit(); s.refresh(ak)
        item = AnswerKeyItem(answer_key_id=ak.id, question_number=1, version="A",
                             correct_answer="", question_type="open_response", weight=1.0)
        s.add(item); s.commit(); s.refresh(item)
        item_id = str(item.id)
    TS = sessionmaker(bind=engine)

    def _override():
        db = TS()
        try:
            yield db
        finally:
            db.close()

    app.dependency_overrides[get_db] = _override
    app.dependency_overrides[usuario_actual] = lambda: _CREADOR
    yield {"item_id": item_id, "engine": engine, "client": TestClient(app)}
    app.dependency_overrides.clear()


def _post(client, item_id, confirmar):
    return client.post(
        f"/api/v1/answer-key-items/{item_id}/rubrica/importar?confirmar={str(confirmar).lower()}",
        files={"file": ("rubrica.xlsx", _xlsx_rubrica(),
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})


def test_import_preview_no_escribe(entorno):
    r = _post(entorno["client"], entorno["item_id"], confirmar=False)
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["guardado"] is False and body["n_criterios"] == 3
    with Session(entorno["engine"]) as s:
        assert s.query(RubricCriterion).count() == 0        # preview NO escribio


def test_import_confirmar_crea_criterios(entorno):
    r = _post(entorno["client"], entorno["item_id"], confirmar=True)
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["guardado"] is True and body["criterios_creados"] == 3
    with Session(entorno["engine"]) as s:
        crits = s.query(RubricCriterion).all()
        assert len(crits) == 3
        integ = next(c for c in crits if c.name == "Integración")
        assert integ.weight == 2.0 and integ.ambito == "grupal"
        assert [n["puntos"] for n in integ.niveles_json] == [3, 2, 1, 0]
