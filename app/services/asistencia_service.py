"""
Motor de asistencia: nómina, sesiones con ventana horaria y QR dinámico firmado.

El QR NO lleva una URL abierta: lleva un DESAFÍO firmado (HMAC del secreto de la sesión
sobre el bucket temporal), que rota cada 4 s. El servidor lo recomputa y acepta solo el
bucket vigente o el inmediatamente anterior (tolerancia de latencia), dentro de la ventana
horaria de la sesión. Así un QR viejo o repetido se rechaza. La marca definitiva se aprueba
con una aserción WebAuthn sobre ese desafío (AS3): passkey = identidad + dispositivo + dueño.

Python puro (hmac/hashlib/time); la cripto de passkeys va en asistencia_webauthn (AS2/AS3).
"""
from __future__ import annotations

import base64
import hashlib
import hmac
import secrets
import time
import uuid as _uuidmod
from datetime import datetime, timezone

from app.core.errors import conflict, not_found
from app.models.asistencia import (
    AsistenciaMatricula, DispositivoWebAuthn, SesionAsistencia, MarcaAsistencia,
    MAT_INVITADO, MAT_VALIDADO, MAT_ACTIVO, SES_ABIERTA, SES_CERRADA,
    MARCA_PRESENTE, MARCA_REVISADO,
)

BUCKET_SEG = 4                     # el QR rota cada 4 s (rápido = anti-compartir, con margen para escanear+firmar)
_TOLERANCIA_SEG = 12               # aceptar la marca hasta ~12 s tarde (latencia móvil + ceremonia passkey)
_ALFABETO = "ABCDEFGHJKLMNPQRSTUVWXYZ23456789"


# ── utilidades de tiempo / firma ─────────────────────────────────────────────────────
def _bucket_actual() -> int:
    return int(time.time() // BUCKET_SEG)


def _digest(secreto: str, sesion_id: str, bucket: int) -> bytes:
    """HMAC-SHA256 (32 bytes) del bucket temporal. Es el desafío que firma la passkey."""
    return hmac.new(secreto.encode(), f"{sesion_id}:{bucket}".encode(), hashlib.sha256).digest()


def _firmar(secreto: str, sesion_id: str, bucket: int) -> str:
    return base64.urlsafe_b64encode(_digest(secreto, sesion_id, bucket)).decode().rstrip("=")[:22]


def desafio_vigente(s, token, bucket) -> bytes | None:
    """Si el desafío del QR (token+bucket) es válido y vigente, devuelve sus 32 bytes
    (el challenge WebAuthn); si no, None. Une la aserción de la passkey a ESE QR fresco."""
    ok, _motivo = verificar_desafio(s, token, bucket)
    if not ok:
        return None
    return _digest(s.secreto, str(s.id), int(bucket))


def _aware(dt) -> datetime:
    """Normaliza a UTC-aware. Acepta datetime o string ISO (sqlite/JSON devuelven variado)."""
    if dt is None:
        return datetime.now(timezone.utc)
    if isinstance(dt, str):
        try:
            dt = datetime.fromisoformat(dt.replace("Z", "+00:00"))
        except ValueError:
            return datetime.now(timezone.utc)
    if dt.tzinfo is None:
        return dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(timezone.utc)


def _uid(x):
    """Coacciona a uuid.UUID (los ids llegan como string desde JSON)."""
    try:
        return x if isinstance(x, _uuidmod.UUID) else _uuidmod.UUID(str(x))
    except (ValueError, TypeError):
        raise not_found("Identificador no válido.")


def _codigo(db) -> str:
    for _ in range(30):
        c = "".join(secrets.choice(_ALFABETO) for _ in range(6))
        if not db.query(SesionAsistencia).filter(SesionAsistencia.codigo == c).first():
            return c
    raise conflict("No se pudo generar un código de sesión único.")


# ── nómina (enrolamiento) ────────────────────────────────────────────────────────────
def parse_nomina_xlsx(data: bytes) -> list[dict]:
    """Extrae filas de un Excel de nómina. Detecta columnas por encabezado (flexible)."""
    import io
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    filas = list(ws.iter_rows(values_only=True))
    if not filas:
        return []
    encab = [str(c or "").strip().lower() for c in filas[0]]

    def _col(*claves):
        for i, h in enumerate(encab):
            if any(k in h for k in claves):
                return i
        return None
    ci = {"nombre": _col("nombre"), "correo": _col("correo", "email", "mail"),
          "identificador": _col("identificador", "id academ", "matricula", "matrícula"),
          "rut": _col("rut", "run", "dni"), "carrera": _col("carrera", "programa"),
          "seccion": _col("sección", "seccion"), "asignatura": _col("asignatura", "ramo", "curso")}
    out = []
    for fila in filas[1:]:
        def g(k):
            j = ci[k]
            return (str(fila[j]).strip() if (j is not None and j < len(fila) and fila[j] is not None) else "")
        correo = g("correo").lower()
        nombre = g("nombre")
        if not (correo or nombre):
            continue
        out.append({"nombre": nombre or correo, "correo": correo, "identificador": g("identificador"),
                    "rut": g("rut"), "carrera": g("carrera"), "seccion": g("seccion"),
                    "asignatura": g("asignatura")})
    return out


def importar_nomina(db, course_id, filas: list[dict]) -> dict:
    """Crea/actualiza matrículas desde la nómina. Genera invite_token para el enrolamiento."""
    creados = actualizados = 0
    for f in (filas or []):
        correo = str(f.get("correo", "")).strip().lower()
        nombre = str(f.get("nombre", "")).strip()
        if not (correo or nombre):
            continue
        m = None
        if correo:
            m = db.query(AsistenciaMatricula).filter(
                AsistenciaMatricula.course_id == course_id,
                AsistenciaMatricula.correo == correo).first()
        if m:
            m.nombre = nombre or m.nombre
            for k in ("identificador", "rut", "carrera", "seccion", "asignatura"):
                if f.get(k):
                    setattr(m, k, str(f[k]).strip())
            actualizados += 1
        else:
            db.add(AsistenciaMatricula(
                course_id=course_id, nombre=nombre or correo, correo=correo or f"sin-correo-{secrets.token_hex(3)}",
                identificador=f.get("identificador") or None, rut=f.get("rut") or None,
                carrera=f.get("carrera") or None, seccion=f.get("seccion") or None,
                asignatura=f.get("asignatura") or None,
                estado=MAT_INVITADO, invite_token=secrets.token_urlsafe(24)))
            creados += 1
    db.commit()
    return {"creados": creados, "actualizados": actualizados}


def listar_nomina(db, course_id) -> list[dict]:
    ms = db.query(AsistenciaMatricula).filter(AsistenciaMatricula.course_id == course_id).all()
    # invite_token se expone SOLO al docente (autorizado) para enviar la invitación de enrolamiento.
    return [{"id": str(m.id), "nombre": m.nombre, "correo": m.correo, "estado": m.estado,
             "identificador": m.identificador, "seccion": m.seccion,
             "invite_token": m.invite_token,
             "tiene_passkey": any(d.activo for d in m.dispositivos)} for m in ms]


def validar_presencial(db, matricula_id) -> dict:
    """El docente confirma presencialmente la identidad -> habilita el registro de passkey."""
    m = db.query(AsistenciaMatricula).filter(AsistenciaMatricula.id == _uid(matricula_id)).first()
    if not m:
        raise not_found("Matrícula no encontrada.")
    if m.estado == MAT_INVITADO:
        m.estado = MAT_VALIDADO
    m.validado_at = datetime.now(timezone.utc)
    db.commit()
    return {"id": str(m.id), "estado": m.estado}


# ── sesiones de asistencia ───────────────────────────────────────────────────────────
def abrir_sesion(db, course_id, abierta_por, titulo, fecha, inicio, fin) -> SesionAsistencia:
    s = SesionAsistencia(course_id=course_id, abierta_por=str(abierta_por),
                         titulo=(titulo or "Asistencia")[:200], fecha=str(fecha)[:10],
                         inicio=_aware(inicio), fin=_aware(fin), estado=SES_ABIERTA,
                         codigo=_codigo(db), secreto=secrets.token_urlsafe(32))
    db.add(s); db.commit(); db.refresh(s)
    return s


def _sesion(db, codigo) -> SesionAsistencia:
    s = db.query(SesionAsistencia).filter(SesionAsistencia.codigo == str(codigo).upper()).first()
    if not s:
        raise not_found("Sesión de asistencia no encontrada.")
    return s


def cerrar_sesion(db, codigo) -> dict:
    s = _sesion(db, codigo)
    s.estado = SES_CERRADA
    db.commit()
    return {"codigo": s.codigo, "estado": s.estado}


def qr_actual(db, codigo) -> dict:
    """Desafío firmado vigente (rota cada 4 s). El frontend lo pinta como QR + cuenta regresiva."""
    s = _sesion(db, codigo)
    b = _bucket_actual()
    return {"codigo": s.codigo, "token": _firmar(s.secreto, str(s.id), b), "bucket": b,
            "vence_en": BUCKET_SEG - int(time.time() % BUCKET_SEG), "rota_cada": BUCKET_SEG,
            "estado": s.estado, "titulo": s.titulo}


def verificar_desafio(s: SesionAsistencia, token: str, bucket) -> tuple[bool, str]:
    """Valida el desafío del QR: firma correcta, bucket vigente o anterior, dentro de ventana."""
    if s.estado != SES_ABIERTA:
        return False, "La sesión de asistencia está cerrada."
    now = datetime.now(timezone.utc)
    if not (_aware(s.inicio) <= now <= _aware(s.fin)):
        return False, "Fuera de la ventana horaria de la asistencia."
    try:
        bucket = int(bucket)
    except (TypeError, ValueError):
        return False, "Desafío inválido."
    actual = _bucket_actual()
    n_atras = max(1, _TOLERANCIA_SEG // BUCKET_SEG)   # nº de buckets hacia atrás que cubren la tolerancia
    for cand in range(actual, actual - n_atras - 1, -1):
        if bucket == cand and hmac.compare_digest(str(token or ""), _firmar(s.secreto, str(s.id), cand)):
            return True, ""
    return False, "El código QR venció; escanea el que está en pantalla."


# ── marcado (núcleo; la aserción passkey se enchufa en AS3) ───────────────────────────
def _anomalias(db, s, matricula_id, ip, ua) -> list:
    flags = []
    # mismo dispositivo (IP+UA) marcando varias matrículas en esta sesión -> señal, no rechazo
    if ip:
        otras = db.query(MarcaAsistencia).filter(
            MarcaAsistencia.sesion_id == s.id, MarcaAsistencia.ip == ip,
            MarcaAsistencia.matricula_id != matricula_id).count()
        if otras:
            flags.append("mismo_dispositivo_multiple")
    return flags


def marcar_verificado(db, s, matricula, ip=None, ua=None, metodo="passkey", flags_extra=None) -> dict:
    """Crea la marca tras la verificación (QR o passkey). Idempotente por (sesión, matrícula).
    Corre antifraude y deja banderas (no rechaza). El docente es la autoridad final."""
    ya = db.query(MarcaAsistencia).filter(
        MarcaAsistencia.sesion_id == s.id, MarcaAsistencia.matricula_id == matricula.id).first()
    if ya:
        return {"estado": ya.estado, "duplicada": True}
    flags = _anomalias(db, s, matricula.id, ip, ua) + list(flags_extra or [])
    marca = MarcaAsistencia(sesion_id=s.id, matricula_id=matricula.id, metodo=metodo,
                            estado=MARCA_REVISADO if flags else MARCA_PRESENTE,
                            anomalias=flags or None, ip=(ip or "")[:64], user_agent=(ua or "")[:300])
    db.add(marca); db.commit()
    return {"estado": marca.estado, "anomalias": flags, "duplicada": False}


def registrar_marca(db, codigo, matricula_id, token, bucket, ip=None, ua=None, metodo="passkey") -> dict:
    """Marca verificando SOLO el desafío del QR (sin passkey). Vía de respaldo/fallback."""
    s = _sesion(db, codigo)
    ok, motivo = verificar_desafio(s, token, bucket)
    if not ok:
        raise conflict(motivo)
    m = db.query(AsistenciaMatricula).filter(
        AsistenciaMatricula.id == _uid(matricula_id), AsistenciaMatricula.course_id == s.course_id).first()
    if not m:
        raise not_found("La matrícula no pertenece a este curso.")
    return marcar_verificado(db, s, m, ip=ip, ua=ua, metodo=metodo)


def estado_sesion(db, codigo) -> dict:
    """Panel docente: nómina, marcas, banderas de anomalía y cobertura."""
    s = _sesion(db, codigo)
    nomina = db.query(AsistenciaMatricula).filter(AsistenciaMatricula.course_id == s.course_id).all()
    marcas = {mk.matricula_id: mk for mk in
              db.query(MarcaAsistencia).filter(MarcaAsistencia.sesion_id == s.id).all()}
    filas = []
    for m in nomina:
        mk = marcas.get(m.id)
        filas.append({"matricula_id": str(m.id), "nombre": m.nombre, "seccion": m.seccion,
                      "presente": bool(mk), "estado": (mk.estado if mk else None),
                      "anomalias": (mk.anomalias if mk else None),
                      "hora": (mk.marcada_at.isoformat() if mk and mk.marcada_at else None)})
    presentes = sum(1 for f in filas if f["presente"])
    return {"codigo": s.codigo, "titulo": s.titulo, "fecha": s.fecha, "estado": s.estado,
            "inicio": _aware(s.inicio).isoformat(), "fin": _aware(s.fin).isoformat(),
            "abierta_at": (_aware(s.created_at).isoformat() if s.created_at else None),
            "abierta_por": s.abierta_por,
            "total": len(nomina), "presentes": presentes, "ausentes": len(nomina) - presentes,
            "con_anomalia": sum(1 for f in filas if f["anomalias"]), "filas": filas}


def informe_payload(db, codigo, formato) -> dict:
    """Payload para exportar la asistencia (docx/pdf = secciones+tablas; xlsx = hojas).
    Registra fecha, ventana horaria, apertura, y por alumno: presente, hora y anomalías."""
    est = estado_sesion(db, codigo)
    titulo = f"Asistencia · {est['titulo']} · {est['fecha']} · Sala {est['codigo']}"
    hora = lambda x: (x[11:16] if x and len(x) >= 16 else "")   # noqa: E731
    filas = [[f["nombre"], f.get("seccion") or "—",
              "Presente" if f["presente"] else "Ausente",
              hora(f.get("hora")) or "—",
              ", ".join(f.get("anomalias") or []) or ""] for f in est["filas"]]
    headers = ["Estudiante", "Sección", "Estado", "Hora de marca", "Anomalías"]
    if formato == "xlsx":
        resumen = {"nombre": "Resumen", "headers": ["Campo", "Valor"], "rows": [
            ["Sesión", est["titulo"]], ["Fecha", est["fecha"]],
            ["Ventana", est["inicio"] + " → " + est["fin"]],
            ["Abierta", est.get("abierta_at") or ""], ["Código", est["codigo"]],
            ["Total", est["total"]], ["Presentes", est["presentes"]],
            ["Ausentes", est["ausentes"]], ["Con anomalía", est["con_anomalia"]]]}
        return {"hojas": [resumen, {"nombre": "Asistencia", "headers": headers, "rows": filas}]}
    resumen_txt = (f"Sesión: {est['titulo']}. Fecha: {est['fecha']}. Ventana: {est['inicio']} → "
                   f"{est['fin']}. Apertura: {est.get('abierta_at') or '—'}. "
                   f"Presentes: {est['presentes']}/{est['total']} · Ausentes: {est['ausentes']} · "
                   f"Con anomalía: {est['con_anomalia']}.")
    return {"titulo": titulo,
            "secciones": [{"heading": "Resumen", "nivel": 1, "texto": resumen_txt},
                          {"heading": None, "nivel": 2,
                           "texto": "Registro con QR dinámico + passkey. La hora es el instante "
                                    "exacto de marca de cada estudiante. Las anomalías son señales "
                                    "para revisión del docente, no rechazos automáticos."}],
            "tablas": [{"titulo": "Detalle de asistencia", "headers": headers, "rows": filas}]}


def override_marca(db, codigo, matricula_id, estado) -> dict:
    """El docente es la autoridad final: fija presente/ausente/revisado manualmente."""
    from app.models.asistencia import MARCA_AUSENTE
    if estado not in (MARCA_PRESENTE, MARCA_REVISADO, MARCA_AUSENTE):
        raise conflict("Estado no válido.")
    s = _sesion(db, codigo)
    mid = _uid(matricula_id)
    mk = db.query(MarcaAsistencia).filter(
        MarcaAsistencia.sesion_id == s.id, MarcaAsistencia.matricula_id == mid).first()
    if not mk:
        mk = MarcaAsistencia(sesion_id=s.id, matricula_id=mid, metodo="docente", estado=estado)
        db.add(mk)
    else:
        mk.estado = estado
    db.commit()
    return {"matricula_id": str(matricula_id), "estado": estado}
