"""Genera plantilla Excel de nómina en memoria."""
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def generate_nomina_template_bytes() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "NOMINA"
    NAVY,TEAL,BGRAY,LGRAY,WHITE,MGRAY = "08101D","0F8B8D","F5F0E8","E4DED4","FFFFFF","94A3B8"
    def hdr(row,col,val,bg,fg="FFFFFF",bold=True,size=10,align="center"):
        c = ws.cell(row=row,column=col,value=val)
        c.font = Font(name="Arial",bold=bold,color=fg,size=size)
        c.fill = PatternFill("solid",fgColor=bg)
        c.alignment = Alignment(horizontal=align,vertical="center",indent=1 if align=="left" else 0)
        thin = Side(style="thin",color=LGRAY)
        c.border = Border(top=thin,bottom=thin,left=thin,right=thin)
    def data(row,col,val="",bg=WHITE,align="left"):
        c = ws.cell(row=row,column=col,value=val)
        c.font = Font(name="Arial",size=10)
        c.fill = PatternFill("solid",fgColor=bg)
        c.alignment = Alignment(horizontal=align,vertical="center",indent=1)
        thin = Side(style="thin",color=LGRAY)
        c.border = Border(top=thin,bottom=thin,left=thin,right=thin)
    ws.merge_cells("A1:F1")
    c=ws["A1"]; c.value="EVALYS — Plantilla de Nómina"
    c.font=Font(name="Arial",bold=True,size=13,color=WHITE)
    c.fill=PatternFill("solid",fgColor=NAVY)
    c.alignment=Alignment(horizontal="center",vertical="center")
    ws.row_dimensions[1].height=28
    ws.merge_cells("A2:F2")
    c=ws["A2"]; c.value="Complete RUT **o** MATRÍCULA (basta uno) y el NOMBRE. Guarde y suba el archivo a Evalys."
    c.font=Font(name="Arial",size=9,italic=True,color=MGRAY)
    c.fill=PatternFill("solid",fgColor=BGRAY)
    c.alignment=Alignment(horizontal="center",vertical="center")
    ws.merge_cells("A3:F3")
    c=ws["A3"]; c.value="⚠  RUT sin puntos, con guion y dígito verificador (ej: 12345678-9). Si el alumno no tiene RUT, escriba su nº de matrícula."
    c.font=Font(name="Arial",size=9,bold=True,color="B45309")
    c.fill=PatternFill("solid",fgColor="FEF9EE")
    c.alignment=Alignment(horizontal="center",vertical="center")
    cols=[("A","N°",6),("B","RUT",20),("C","MATRÍCULA",18),("D","APELLIDO PATERNO *",24),("E","APELLIDO MATERNO *",24),("F","NOMBRES *",28)]
    for col_letter,label,width in cols:
        col_idx=ord(col_letter)-ord("A")+1
        hdr(4,col_idx,label,TEAL)
        ws.column_dimensions[col_letter].width=width
    ws.row_dimensions[4].height=24
    samples=[(1,"12345678-9","","González","Muñoz","Catalina Andrea"),
             (2,"98765432-1","","Figueroa","Soto","Valentina Paz"),
             (3,"","A2026-1187","Pérez","Da Silva","João Andrés")]
    for i,row_data in enumerate(samples):
        r=5+i; bg=BGRAY if i%2==0 else WHITE
        for j,val in enumerate(row_data):
            data(r,j+1,val,bg=bg,align="center" if j==0 else "left")
        ws.row_dimensions[r].height=18
    for i in range(3,100):
        r=5+i; bg=BGRAY if i%2==0 else WHITE
        data(r,1,i+1,bg=bg,align="center")
        for j in range(2,7): data(r,j,bg=bg)
        ws.row_dimensions[r].height=18
    ws.freeze_panes="A5"
    buf=io.BytesIO(); wb.save(buf); return buf.getvalue()
