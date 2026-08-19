import re, io


def _dv_ok(num_str, dv):
    """Verifica el dígito verificador chileno (módulo 11)."""
    total = sum(int(d) * [2, 3, 4, 5, 6, 7, 2, 3][i] for i, d in enumerate(reversed(num_str)))
    r = 11 - (total % 11)
    exp = "0" if r == 11 else "K" if r == 10 else str(r)
    return dv == exp


def clean_rut(raw):
    """Normaliza un RUT/RUN a 'num-dv'. Devuelve (normalizado|None, dv_ok).
    Acepta el RUT aunque el DV no calce (dv_ok=False) — NO lo descarta: muchas nóminas reales
    traen DV imperfectos y el docente igual necesita la lista. Devuelve None solo si no parece RUT."""
    if raw is None:
        return None, False
    s = str(raw).strip().upper().replace(".", "").replace(" ", "")
    # Insertar guión si viene pegado (12345678K -> 12345678-K)
    if "-" not in s and len(s) >= 2 and s[-1] in "0123456789K":
        s = s[:-1] + "-" + s[-1]
    m = re.match(r"^(\d{6,9})-([0-9K])$", s)
    if not m:
        return None, False
    num_str, dv = m.group(1), m.group(2)
    return f"{num_str}-{dv}", _dv_ok(num_str, dv)


def validate_rut(rut):
    """Compat: (checksum_ok, normalizado)."""
    norm, ok = clean_rut(rut)
    return (bool(norm) and ok), (norm or str(rut or "").strip())


def _norm(v):
    return str(v or "").strip().upper()


def _looks_like_rut(v):
    return clean_rut(v)[0] is not None


# Encabezados que identifican la columna de MATRÍCULA / identificador académico (nóminas sin RUT:
# extranjeros, universidades que no usan RUN, intercambio…).
_MAT_HEADERS = ("MATRICULA", "MATRÍCULA", "MATRICULA N", "N MATRICULA", "NRO MATRICULA",
                "NUMERO DE MATRICULA", "N° ALUMNO", "NRO ALUMNO", "ID ALUMNO", "ID ESTUDIANTE",
                "CODIGO ALUMNO", "CÓDIGO ALUMNO", "CODIGO ESTUDIANTE", "REGISTRO", "CARNET",
                "LEGAJO", "IDENTIFICADOR", "PASAPORTE")


def _es_col_matricula(v):
    v = _norm(v)
    if not v or len(v) > 28:
        return False
    return any(h in v for h in _MAT_HEADERS)


def clean_matricula(raw):
    """Normaliza un nº de matrícula/identificador: sin espacios ni puntos, en mayúsculas."""
    s = str(raw or "").strip().upper().replace(" ", "").replace(".", "")
    return s if len(s) >= 3 else ""


def parse_nomina_excel(file_bytes):
    try:
        import openpyxl
    except Exception:
        return {"error": "openpyxl no disponible", "students": [], "errors": []}
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as e:
        return {"error": "No se pudo abrir el Excel: " + str(e), "students": [], "errors": []}

    sheet = next((wb[n] for n in wb.sheetnames if "NOMINA" in n.upper() or "NÓMINA" in n.upper()),
                 wb.active)
    MAXC = min(sheet.max_column or 20, 30)

    # ── 1. Detectar la fila de encabezado (tolerante a variantes de nombre de columna) ──
    hrow = rcol = apcol = amcol = ncol = apellidos_col = matcol = None
    for ri in range(1, min((sheet.max_row or 20), 20) + 1):
        vals = {ci: _norm(sheet.cell(ri, ci).value) for ci in range(1, MAXC + 1)}
        rut_cols = [ci for ci, v in vals.items() if ("RUT" in v or "RUN" in v) and len(v) <= 24]
        mat_cols = [ci for ci, v in vals.items() if _es_col_matricula(v)]
        if not rut_cols and not mat_cols:          # sirve RUT **o** matrícula
            continue
        pat = [ci for ci, v in vals.items() if "PATERNO" in v]
        mat = [ci for ci, v in vals.items() if "MATERNO" in v]
        nom = [ci for ci, v in vals.items() if "NOMBRE" in v]
        # "APELLIDOS" (columna combinada, sin paterno/materno separados)
        apes = [ci for ci, v in vals.items() if "APELLIDO" in v and "PATERNO" not in v and "MATERNO" not in v]
        if pat or nom or apes:
            hrow = ri
            rcol = rut_cols[0] if rut_cols else None
            matcol = mat_cols[0] if mat_cols else None
            apcol = pat[0] if pat else None
            amcol = mat[0] if mat else None
            ncol = nom[0] if nom else None
            apellidos_col = apes[0] if (apes and not pat) else None
            break

    modo_posicional = False
    if not hrow:
        # ── Fallback SIN encabezado: si la 1ª celda de alguna de las primeras filas es un RUT,
        # asumimos col A=RUT, B=apellidos, C=nombres (o B=nombre completo). ──
        for ri in range(1, min((sheet.max_row or 5), 5) + 1):
            if _looks_like_rut(sheet.cell(ri, 1).value):
                hrow = ri - 1     # los datos empiezan en esta fila
                rcol, apellidos_col, ncol = 1, 2, 3
                modo_posicional = True
                break
    if not hrow and not modo_posicional:
        return {"error": "No se reconocieron columnas de identificación y nombre en el Excel. "
                         "Asegúrate de tener encabezados como 'RUT' (o 'Matrícula'), "
                         "'Apellido Paterno', 'Apellido Materno', 'Nombres' (o usa la plantilla).",
                "students": [], "errors": [], "valid_count": 0, "error_count": 0, "total": 0}

    students, errors, seen = [], [], set()
    dv_warn = 0
    sin_rut = 0
    for ri in range(hrow + 1, (sheet.max_row or hrow) + 1):
        raw = sheet.cell(ri, rcol).value if rcol else None
        raw_mat = sheet.cell(ri, matcol).value if matcol else None
        if (raw is None or str(raw).strip() == "") and (raw_mat is None or str(raw_mat).strip() == ""):
            continue
        # Filas-resumen al pie de las planillas de notas (no son alumnos): se ignoran en silencio.
        _low = re.sub(r"[^a-z]", "", str(raw if raw is not None else raw_mat).strip().lower())
        if _low in ("promedio", "promedios", "media", "desviacion", "desviacionestandar", "desv",
                    "total", "totales", "maximo", "minimo", "mediana", "moda", "aprobados", "reprobados"):
            continue
        norm, dvok = clean_rut(raw) if raw is not None else (None, False)
        matricula = clean_matricula(raw_mat)
        # Si no hay columna de matrícula pero el valor de la columna RUT no es un RUT, puede ser
        # una matrícula escrita ahí (nóminas mixtas): se acepta como identificador.
        if not norm and not matricula and raw is not None:
            matricula = clean_matricula(raw)
        ap = str(sheet.cell(ri, apcol).value or "").strip() if apcol else ""
        am = str(sheet.cell(ri, amcol).value or "").strip() if amcol else ""
        nm = str(sheet.cell(ri, ncol).value or "").strip() if ncol else ""
        if apellidos_col and not ap:          # columna combinada "Apellidos"
            ap = str(sheet.cell(ri, apellidos_col).value or "").strip()
        name = " ".join(x for x in (ap, am, nm) if x).strip() or ("Estudiante " + str(ri - hrow))
        # Identificador del alumno: el RUT si es válido; si no, la matrícula. Se propaga como `rut`
        # para que TODO el pipeline aguas abajo (escaneos, matriz, libro de notas) siga calzando.
        ident = norm or matricula
        if not ident:
            errors.append({"row": ri, "rut": str(raw or raw_mat or ""), "name": name,
                           "error": "Sin RUT válido ni número de matrícula"})
            continue
        if ident in seen:
            errors.append({"row": ri, "rut": ident, "name": name,
                           "error": ("RUT duplicado (se omitió)" if norm else "Matrícula duplicada (se omitió)")})
            continue
        seen.add(ident)
        if norm and not dvok:
            dv_warn += 1
        if not norm:
            sin_rut += 1
        students.append({"rut": ident, "matricula": matricula or None, "tiene_rut": bool(norm),
                         "name": name, "apellido_paterno": ap,
                         "apellido_materno": am, "nombres": nm, "dv_ok": dvok})

    return {
        "students": students,
        "errors": errors,
        "total": len(students) + len(errors),
        "valid_count": len(students),
        "error_count": len(errors),
        "dv_advertencias": dv_warn,
        "sin_rut": sin_rut,          # alumnos identificados por matrícula (sin RUT)
    }
