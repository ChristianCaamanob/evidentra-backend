"""
Import de rubrica desde .xlsx (para que el docente no re-tipee su rubrica).

Parsea la estructura de rubrica ANALITICA comun -fila = criterio, columnas = niveles con
puntaje, una columna de factor, encabezados de seccion- como la de 'Presentacion Infografia
DMOR 0030'. Es HEURISTICO y tolerante, no universal: devuelve un preview estructurado que el
docente confirma antes de guardar (compuerta humana).

Detecta:
  - la fila de encabezado (celda 'Categoria' + >=2 niveles con formato 'Nombre (n puntos)').
  - la columna de factor de multiplicacion ('Factor...').
  - filas de SECCION (categoria sin descriptores de nivel; 'Grupal'/'Individual' fija el ambito).
  - filas de CRITERIO (categoria + descriptores por nivel + factor).

Salida: {criterios:[{name, seccion, ambito, weight, niveles:[{nivel,puntos,descriptor}]}], ...}
"""
from __future__ import annotations

import io
import re
import unicodedata

from app.core.errors import unprocessable


def _norm(s) -> str:
    s = unicodedata.normalize("NFKD", str(s or "")).encode("ascii", "ignore").decode()
    return s.lower().strip()


_NIVEL_RE = re.compile(r"(.+?)\(\s*(\d+)\s*puntos?\s*\)", re.I)


def _num_a_la_derecha(ws, r, c, maxc):
    """Primer valor numérico en las celdas a la derecha de (r,c)."""
    for cc in range(c + 1, min(maxc, c + 5) + 1):
        v = ws.cell(r, cc).value
        if isinstance(v, (int, float)):
            return float(v)
        if v is not None:
            try:
                return float(str(v).replace("%", "").replace(",", ".").strip())
            except (TypeError, ValueError):
                continue
    return None


def _detectar_config(ws, maxr, maxc):
    """Detecta los parámetros de calificación de la planilla (exigencia, nota mín/máx,
    puntaje máx) para configurar la evaluación sin re-tipear. Heurístico y tolerante."""
    cfg = {}
    for r in range(1, maxr + 1):
        for c in range(1, maxc + 1):
            v = ws.cell(r, c).value
            if v is None or isinstance(v, (int, float)):
                continue
            t = _norm(v)
            val = _num_a_la_derecha(ws, r, c, maxc)
            if val is None:
                continue
            if ("para 4" in t or "exigencia" in t) and 0 < val <= 100 and "exigencia" not in cfg:
                cfg["exigencia"] = round(val, 1)
            elif ("nota minima" in t or "nota min" in t) and val <= 10 and "nota_min" not in cfg:
                cfg["nota_min"] = val
            elif ("nota maxima" in t or "nota max" in t) and val <= 10 and "nota_max" not in cfg:
                cfg["nota_max"] = val
            elif ("puntaje max" in t or "ptje max" in t) and "puntaje_max" not in cfg:
                cfg["puntaje_max"] = val
    # Escala sugerida por la nota máxima (o chilena por defecto si el piso es ~1).
    nm = cfg.get("nota_max")
    if nm and abs(nm - 7) < 0.6:
        cfg["escala"] = "chile_1_7"
    elif nm and abs(nm - 10) < 0.6:
        cfg["escala"] = "europe_10"
    elif nm and abs(nm - 20) < 0.9:
        cfg["escala"] = "francia_20"
    elif "escala" not in cfg and cfg.get("nota_min", 1) <= 1.5:
        cfg["escala"] = "chile_1_7"     # piso 1.0 -> escala chilena
    return cfg or None


def parse_rubrica_xlsx(data: bytes) -> dict:
    import openpyxl

    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    except Exception as e:
        raise unprocessable(f"No se pudo abrir el archivo como Excel: {e}")
    ws = wb.worksheets[0]
    maxr, maxc = min(ws.max_row, 300), min(ws.max_column, 30)

    # --- encabezado: 'Categoria' + niveles 'Nombre (n puntos)' ---
    header_row = cat_col = factor_col = None
    nivel_cols = []
    for r in range(1, maxr + 1):
        textos = {c: ws.cell(r, c).value for c in range(1, maxc + 1)}
        textos = {c: str(v).strip() for c, v in textos.items() if v is not None}
        cat_c = next((c for c, t in textos.items() if _norm(t).startswith("categoria")), None)
        niveles = []
        for c, t in textos.items():
            m = _NIVEL_RE.search(t.replace("\n", " "))
            if m:
                niveles.append((c, m.group(1).strip(), int(m.group(2))))
        if cat_c and len(niveles) >= 2:
            header_row, cat_col = r, cat_c
            nivel_cols = sorted(niveles)
            factor_col = next((c for c, t in textos.items() if "factor" in _norm(t)), None)
            break

    if header_row is None:
        raise unprocessable(
            "No se reconocio una tabla de rubrica: falta un encabezado con 'Categoria' y "
            "niveles con formato 'Nombre (n puntos)'.")

    # --- seccion inicial: mirar por encima del encabezado ---
    seccion, ambito = None, "individual"

    def _ambito_de(txt):
        if re.search(r"grupal", txt, re.I):
            return "grupal"
        if re.search(r"individual", txt, re.I):
            return "individual"
        return None

    for r in range(header_row - 1, 0, -1):
        v = next((ws.cell(r, c).value for c in range(1, maxc + 1) if ws.cell(r, c).value), None)
        if v and re.search(r"puntaje\s+(grupal|individual)|presentaci", str(v), re.I):
            seccion = re.sub(r"\(.*?\)", "", str(v)).strip() or str(v).strip()
            ambito = _ambito_de(str(v)) or ambito
            break

    # --- filas de criterio / seccion ---
    criterios, advertencias = [], []
    for r in range(header_row + 1, maxr + 1):
        cat = ws.cell(r, cat_col).value
        if cat is None:
            continue
        cat = str(cat).strip()
        descriptores = {}
        for c, nivel, pts in nivel_cols:
            v = ws.cell(r, c).value
            if v is not None and str(v).strip():
                descriptores[(nivel, pts)] = str(v).strip()
        if not descriptores:                        # fila de SECCION
            seccion = re.sub(r"\(.*?\)", "", cat).strip() or cat
            ambito = _ambito_de(cat) or ambito
            continue
        # Descartar filas espurias (p. ej. la tabla de conversion puntaje->nota del final):
        # un criterio real tiene la MAYORIA de los niveles con descriptor y un nombre no numerico.
        if len(descriptores) < max(2, len(nivel_cols) - 1) or re.fullmatch(r"[\d.,\s]+", cat):
            continue
        factor = 1.0
        if factor_col is not None and ws.cell(r, factor_col).value is not None:
            try:
                factor = float(ws.cell(r, factor_col).value)
            except (TypeError, ValueError):
                advertencias.append(f"Factor no numerico en '{cat}'; se uso 1.")
        niveles = [{"nivel": nivel, "puntos": pts,
                    "descriptor": descriptores.get((nivel, pts), "")}
                   for _, nivel, pts in nivel_cols]
        criterios.append({"name": cat, "seccion": seccion, "ambito": ambito,
                          "weight": factor, "niveles": niveles})

    if not criterios:
        raise unprocessable("Se reconocio el encabezado pero ningun criterio con descriptores.")

    return {
        "criterios": criterios,
        "n_criterios": len(criterios),
        "escala": [{"nivel": n, "puntos": p} for _, n, p in nivel_cols],
        "secciones": sorted({c["seccion"] for c in criterios if c["seccion"]}),
        "config": _detectar_config(ws, maxr, maxc),   # exigencia/escala detectadas de la planilla
        "advertencias": advertencias,
        "nota": "Preview heuristico: revisa y confirma antes de guardar (el docente decide, G1).",
    }
