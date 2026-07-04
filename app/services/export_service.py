"""
P2 - Exportacion del informe del curso y del dataset de investigacion.

Formatos:
  - CSV  : dataset 'largo' (tidy) UTF-8 -> listo para R / Python / SPSS / jamovi.
  - XLSX : libro multi-hoja (Resumen, Items, Nomina, Datos_largo).
  - PDF  : informe formateado del curso (resumen + items + por RA).

Se alimenta del resultado de curso_stats_service.analizar_evaluacion(...).
"""
from __future__ import annotations

import csv
import io


# ───────────────────────────────────────────────────────────── CSV (tidy)
def dataset_csv(resultado: dict) -> str:
    """Devuelve el dataset 'largo' como texto CSV (una fila por alumno-item)."""
    filas = resultado.get("dataset_largo", [])
    campos = ["student_id", "item", "ra", "bloom", "unidad", "elegida", "correcta", "correcto"]
    buf = io.StringIO()
    w = csv.DictWriter(buf, fieldnames=campos, extrasaction="ignore")
    w.writeheader()
    for f in filas:
        w.writerow(f)
    return buf.getvalue()


def to_csv(resultado: dict, path: str) -> str:
    with open(path, "w", encoding="utf-8", newline="") as fh:
        fh.write(dataset_csv(resultado))
    return path


# ───────────────────────────────────────────────────────────── XLSX
def to_xlsx(resultado: dict, path: str) -> str:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    bold = Font(bold=True)

    # Hoja Resumen
    ws = wb.active
    ws.title = "Resumen"
    dn = resultado.get("descriptivos_nota", {})
    dp = resultado.get("descriptivos_pct", {})
    nm = resultado.get("normalidad_nota", {})
    resumen = [
        ("Metrica", "Valor"),
        ("N alumnos", resultado["instrumento"]["n_alumnos"]),
        ("N items", resultado["instrumento"]["n_items"]),
        ("Nota media", dn.get("media")),
        ("Nota mediana", dn.get("mediana")),
        ("Nota DE", dn.get("de")),
        ("IC95 media (nota)", str(dn.get("ic95_media"))),
        ("Logro medio %", dp.get("media")),
        ("Logro DE %", dp.get("de")),
        ("Asimetria (%)", dp.get("asimetria")),
        ("Curtosis exceso (%)", dp.get("curtosis_exceso")),
        ("Tasa aprobacion %", resultado.get("tasa_aprobacion")),
        ("Aprobados", resultado.get("aprobados")),
        ("Reprobados", resultado.get("reprobados")),
        ("Confiabilidad KR-20", resultado.get("confiabilidad_kr20")),
        ("Shapiro-Wilk W", nm.get("W")),
        ("Shapiro-Wilk p", nm.get("p")),
        ("Distribucion", nm.get("distribucion")),
        ("Test recomendado", nm.get("recomendacion_test")),
    ]
    for r in resumen:
        ws.append(r)
    ws["A1"].font = bold; ws["B1"].font = bold
    ws.column_dimensions["A"].width = 26; ws.column_dimensions["B"].width = 40

    # Hoja Items
    wi = wb.create_sheet("Items")
    wi.append(["item", "ra", "bloom", "unidad", "dificultad_p", "nivel",
               "discriminacion_pbis", "calidad", "D_kelley",
               "n_correctas", "n_incorrectas", "n_vacias",
               "distractor_A_%", "distractor_B_%", "distractor_C_%", "distractor_D_%", "distractor_E_%"])
    for c in wi[1]:
        c.font = bold
    for it in resultado.get("items", []):
        d = it["distractores"]
        wi.append([it["item"], it["ra"], it["bloom"], it["unidad"], it["dificultad_p"],
                   it["nivel_dificultad"], it["discriminacion_pbis"], it["calidad_discriminacion"],
                   it["discriminacion_d_kelley"], it["n_correctas"], it["n_incorrectas"], it["n_vacias"],
                   d["A"]["pct"], d["B"]["pct"], d["C"]["pct"], d["D"]["pct"], d["E"]["pct"]])

    # Hoja Nomina
    wn = wb.create_sheet("Nomina")
    wn.append(["student_id", "puntaje", "pct", "nota", "aprobado"])
    for c in wn[1]:
        c.font = bold
    for a in resultado.get("dataset_ancho", []):
        wn.append([a["student_id"], a["puntaje"], a["pct"], a["nota"], a["aprobado"]])

    # Hoja Datos_largo (tidy para investigacion)
    wl = wb.create_sheet("Datos_largo")
    campos = ["student_id", "item", "ra", "bloom", "unidad", "elegida", "correcta", "correcto"]
    wl.append(campos)
    for c in wl[1]:
        c.font = bold
    for f in resultado.get("dataset_largo", []):
        wl.append([f.get(k) for k in campos])

    wb.save(path)
    return path


# ───────────────────────────────────────────────────────────── PDF
def to_pdf(resultado: dict, path: str, titulo: str = "Informe del curso") -> str:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle)
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    styles = getSampleStyleSheet()
    h = ParagraphStyle("h", parent=styles["Heading2"], textColor=colors.HexColor("#0B6E70"))
    small = ParagraphStyle("s", parent=styles["Normal"], fontSize=8, textColor=colors.HexColor("#5E6F72"))

    doc = SimpleDocTemplate(path, pagesize=A4, topMargin=18*mm, bottomMargin=16*mm,
                            leftMargin=18*mm, rightMargin=18*mm, title=titulo)
    E = []
    dn = resultado.get("descriptivos_nota", {})
    nm = resultado.get("normalidad_nota", {})
    E.append(Paragraph(titulo, styles["Title"]))
    E.append(Paragraph(f"n={resultado['instrumento']['n_alumnos']} &middot; "
                       f"{resultado['instrumento']['n_items']} items", small))
    E.append(Spacer(1, 8))

    resumen = [
        ["Metrica", "Valor"],
        ["Nota media / mediana / DE", f"{dn.get('media')} / {dn.get('mediana')} / {dn.get('de')}"],
        ["Tasa de aprobacion", f"{resultado.get('tasa_aprobacion')}%  ({resultado.get('aprobados')}/{resultado['instrumento']['n_alumnos']})"],
        ["Confiabilidad KR-20", str(resultado.get("confiabilidad_kr20"))],
        ["Normalidad (Shapiro-Wilk)", f"W={nm.get('W')}, p={nm.get('p')} -> {nm.get('distribucion')}"],
        ["Test recomendado", nm.get("recomendacion_test", "-")],
    ]
    t = Table(resumen, colWidths=[60*mm, 105*mm])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F8B8D")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#E4EBEB")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F7F7")]),
    ]))
    E.append(t)
    E.append(Spacer(1, 12))

    E.append(Paragraph("Analisis de items", h))
    filas = [["Item", "RA", "p", "Nivel", "Discrim.", "Calidad", "Correcta"]]
    for it in resultado.get("items", []):
        filas.append([f"P{it['item']}", it["ra"], f"{it['dificultad_p']:.2f}", it["nivel_dificultad"],
                      f"{it['discriminacion_pbis']:.2f}", it["calidad_discriminacion"],
                      next(k for k, v in it["distractores"].items() if v["correcta"])])
    ti = Table(filas, repeatRows=1, colWidths=[16*mm, 16*mm, 16*mm, 22*mm, 22*mm, 26*mm, 20*mm])
    ti.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#16252E")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 7.5),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#E4EBEB")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F7FAFA")]),
    ]))
    E.append(ti)
    E.append(Spacer(1, 10))
    if resultado.get("por_ra"):
        E.append(Paragraph("Cobertura por RA (segun la TE de este instrumento)", h))
        fr = [["RA", "Items", "Logro %"]] + [[g["clave"], str(g["items"]), f"{g['logro_promedio']}%"]
                                             for g in resultado["por_ra"]]
        tr = Table(fr, colWidths=[40*mm, 30*mm, 40*mm])
        tr.setStyle(TableStyle([("FONTSIZE", (0, 0), (-1, -1), 8),
                                ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#E4EBEB")),
                                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F2F7F7"))]))
        E.append(tr)
    E.append(Spacer(1, 12))
    E.append(Paragraph("Generado por Evalys — psicometria clasica + normalidad. Documento pedagogico; "
                       "la calificacion es responsabilidad del docente.", small))
    doc.build(E)
    return path
