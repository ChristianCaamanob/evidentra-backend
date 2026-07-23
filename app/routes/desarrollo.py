"""
Router del motor de desarrollo (Fase 2 del cableado): validacion docente (F3) que persiste
la trazabilidad, y los analisis de rubrica que la consumen (R, MFRM) mas las propuestas de
aprendizaje (F4).

  POST /results/{scan_id}/validar                 -> F3: persiste RegistroValidacion (G1, G5)
  GET  /assessments/{id}/rubrica/mfrm             -> I6: severidad IA vs docente
  GET  /assessments/{id}/rubrica/psicometria      -> R:  psicometria de la rubrica + G-theory
  GET  /assessments/{id}/rubrica/aprendizaje      -> F4: propuestas de ajuste (solo propone)

La nota final es del docente (G1); cada decision queda con sello temporal inmutable (G5);
los datos se agregan seudonimizados (G2).
"""
import logging
import traceback
from uuid import UUID

from fastapi import APIRouter, Depends, UploadFile, File
from sqlalchemy.orm import Session

from app.api.deps import get_db, req_profesor, req_investigador
from app.core.errors import not_found, conflict
from app.models.validacion import RegistroValidacion
from app.models.aprendizaje import RubricaVersion, AjusteCalibracion
from app.services import matriz_service
from app.services import validacion_service
from app.services import precalificacion_service
from app.services import coder_llm
from app.services import correccion_experta_service
from app.services import retroalimentacion_service
from app.services import reportes_desarrollo_service
from app.services import mfrm_service
from app.services import rubrica_psicometria_service
from app.services import aprendizaje_service

router = APIRouter(tags=["desarrollo"])
logger = logging.getLogger("evalys")


_persistir_validaciones = validacion_service.persistir_validaciones


@router.post("/results/{scan_id}/validar", dependencies=[Depends(req_profesor)])
def validar_desarrollo(scan_id: UUID, payload: dict, db: Session = Depends(get_db)):
    """
    F3 (escrita) - Valida la respuesta de desarrollo de un ESCANEO. La nota es del docente (G1).

    payload = {docente, rubrica_version_hash?, criterios:[
                 {criterio, nivel_ia?, confianza_ia?, nivel_docente, comentario?}]}
    """
    try:
        scan = matriz_service.scan_repo.get(db, scan_id)
        if not scan:
            raise not_found("Escaneo no encontrado.")
        pseudo = matriz_service._pseudo(scan.id)
        registros, version_hash = _persistir_validaciones(db, pseudo, scan.assessment_id, payload)
        return {
            "sujeto": pseudo, "n_registrados": len(registros),
            "acuerdo": validacion_service.acuerdo_qwk(registros) if registros else {},
            "rubrica_version_hash": version_hash,
            "gobernanza": "Nota fijada por el docente (G1). Trazabilidad inmutable (G5). "
                          "Seudonimizado (G2).",
        }
    except KeyError as e:
        raise not_found(f"Falta el campo {e} en un criterio del payload.")
    except Exception:
        logger.error(f"Error en validar_desarrollo {scan_id}: {traceback.format_exc()}")
        raise


@router.post("/assessments/{assessment_id}/students/{student_id}/rubrica/validar",
             dependencies=[Depends(req_profesor)])
def validar_oral(assessment_id: UUID, student_id: UUID, payload: dict,
                 db: Session = Depends(get_db)):
    """
    F3 (oral) - Aplica la rubrica parametrizada DIRECTAMENTE a un estudiante (oral, presentacion,
    practica): no hay hoja/escaneo, el sujeto es el estudiante de la nomina. Misma trazabilidad
    y gobernanza que la escrita. La IA es opcional (el docente puede puntuar directo).
    """
    try:
        from app.models.student import Student
        from app.models.assessment import Assessment
        st = db.get(Student, student_id)
        if not st:
            raise not_found("Estudiante no encontrado.")
        a = db.get(Assessment, assessment_id)
        if not a:
            raise not_found("Evaluacion no encontrada.")
        if st.course_id != a.course_id:
            raise conflict("El estudiante no pertenece al curso de la evaluacion.")
        pseudo = matriz_service._pseudo(st.id)
        registros, version_hash = _persistir_validaciones(db, pseudo, assessment_id, payload)
        return {
            "sujeto": pseudo, "modalidad": "oral", "n_registrados": len(registros),
            "acuerdo": validacion_service.acuerdo_qwk(registros) if registros else {},
            "rubrica_version_hash": version_hash,
            "gobernanza": "Nota fijada por el docente (G1). Rubrica aplicada directo al estudiante. "
                          "Trazabilidad inmutable (G5). Seudonimizado (G2).",
        }
    except KeyError as e:
        raise not_found(f"Falta el campo {e} en un criterio del payload.")
    except Exception:
        logger.error(f"Error en validar_oral {assessment_id}/{student_id}: {traceback.format_exc()}")
        raise


@router.post("/answer-key-items/{item_id}/transcribir", dependencies=[Depends(req_profesor)])
async def transcribir_respuesta(item_id: UUID, file: UploadFile = File(...), db: Session = Depends(get_db)):
    """Transcribe con IA de visión una respuesta MANUSCRITA (foto/PDF) de una pregunta de
    desarrollo. Devuelve el texto tal cual para que el docente lo revise y pre-califique (G1).
    No corrige ni califica; marca lo ilegible."""
    from app.services import transcripcion_service
    from app.models.answer_key import AnswerKeyItem
    try:
        it = db.get(AnswerKeyItem, item_id)
        enun = (getattr(it, "enunciado", "") or "") if it else ""
        data = await file.read()
        mt = file.content_type or "image/jpeg"
        return transcripcion_service.transcribir(data, mt, enun)
    except Exception:
        logger.error(f"Error en transcribir_respuesta {item_id}: {traceback.format_exc()}")
        raise


@router.post("/answer-key-items/{item_id}/precalificar", dependencies=[Depends(req_profesor)])
def precalificar(item_id: UUID, payload: dict, db: Session = Depends(get_db)):
    """
    F2 - Pre-califica una respuesta de desarrollo criterio por criterio (la IA propone; la
    nota es del docente, G1). Devuelve el hash de la version de rubrica ACTIVA para que el
    docente lo fije al validar (pinning -> replicabilidad).

    payload = {respuesta}
    """
    try:
        criterios = matriz_service.cargar_criterios_item(db, item_id)
        if not criterios:
            raise conflict("El item no tiene criterios de rubrica definidos.")
        from app.models.answer_key import AnswerKeyItem, AnswerKey
        from app.models.assessment import Assessment
        from app.models.course import Course
        item = db.get(AnswerKeyItem, item_id)

        # Norma terminologica heredada del curso (para el modo estricto del LLM).
        ak = db.get(AnswerKey, item.answer_key_id)
        assessment = db.get(Assessment, ak.assessment_id) if ak else None
        course = db.get(Course, assessment.course_id) if assessment else None
        norma = getattr(course, "norma_terminologica", None) if course else None
        for c in criterios:
            c["norma_terminologica"] = norma

        coder = coder_llm.coder_por_defecto()      # LLM si hay API key; si no, grader determinista
        rep = precalificacion_service.precalificar_respuesta(
            payload.get("respuesta", ""), criterios, coder=coder, norma_terminologica=norma)
        # Motor REAL: 'llm' solo si el modelo respondió; si cayó al determinista lo decimos.
        est = getattr(coder, "estado", None) if coder else None
        if coder is None:
            rep["motor"] = "deterministico"
        elif est and est["llm_ok"] > 0 and est["fallback"] == 0:
            rep["motor"] = "llm"
        elif est and est["llm_ok"] > 0:
            rep["motor"] = "mixto"
        else:
            rep["motor"] = "llm_fallback_deterministico"
            if est and est.get("ultimo_error"):
                rep["motor_detalle"] = est["ultimo_error"]
        rep["rubrica_version_hash"] = matriz_service.version_activa_hash(
            db, item.answer_key_id, criterios)
        return rep
    except Exception:
        logger.error(f"Error en precalificar {item_id}: {traceback.format_exc()}")
        raise


@router.post("/answer-key-items/{item_id}/corregir-experto", dependencies=[Depends(req_profesor)])
def corregir_experto(item_id: UUID, payload: dict, db: Session = Depends(get_db)):
    """
    Fase 3 · Corrección EXPERTA holística (la IA PROPONE; el docente valida, G1). Funciona
    AUNQUE no haya rúbrica manual: se apoya en respuesta óptima + nivel de rigor + autoridad
    del área + la fuente que el docente declara (clave en Derecho/normativa, sin alucinar).

    payload = {respuesta}
    """
    from app.models.answer_key import AnswerKeyItem
    it = db.get(AnswerKeyItem, item_id)
    if not it:
        raise not_found("Pregunta no encontrada.")
    respuesta = (payload.get("respuesta") or "").strip()
    if not respuesta:
        raise conflict("Falta la respuesta del estudiante a corregir.")
    criterios = [{"name": c.name, "descriptor": c.descriptor} for c in (it.rubric_criteria or [])]
    cfg = {
        "enunciado": it.enunciado or "",
        "respuesta_optima": (it.respuesta_optima or it.correct_answer or ""),
        "nivel_rigor": getattr(it, "nivel_rigor", None) or "estricto",
        "area_conocimiento": getattr(it, "area_conocimiento", None) or "general",
        "fuente_estandar": getattr(it, "fuente_estandar", None) or "",
        "peso": float(it.weight or 1),
        "criterios": criterios,
    }
    try:
        return correccion_experta_service.corregir(respuesta, cfg)
    except Exception:
        logger.error(f"Error en corregir_experto {item_id}: {traceback.format_exc()}")
        raise


@router.post("/assessments/{assessment_id}/desarrollo/informe-retroalimentacion",
             dependencies=[Depends(req_profesor)])
def informe_retroalimentacion(assessment_id: UUID, payload: dict, db: Session = Depends(get_db)):
    """
    Fase 4 · Informe de retroalimentación que CRUZA la Tabla de Especificaciones (RA/unidad)
    con el desempeño del estudiante → brechas + estrategias de estudio basadas en evidencia
    POR RA. La IA propone; el docente valida (G1). payload = {respuestas:[{item_id,respuesta}],
    estudiante?}.
    """
    from app.models.assessment import Assessment
    asm = db.get(Assessment, assessment_id)
    if not asm:
        raise not_found("Evaluación no encontrada.")
    respuestas = payload.get("respuestas") or []
    if not isinstance(respuestas, list) or not respuestas:
        raise conflict("Envía al menos una respuesta {item_id, respuesta}.")
    estudiante = (str(payload.get("estudiante") or "Estudiante").strip() or "Estudiante")[:80]
    try:
        return retroalimentacion_service.generar_informe(db, asm, respuestas, estudiante)
    except Exception:
        logger.error(f"Error en informe_retroalimentacion {assessment_id}: {traceback.format_exc()}")
        raise


@router.get("/assessments/{assessment_id}/desarrollo/reportes", dependencies=[Depends(req_profesor)])
def reportes_desarrollo(assessment_id: UUID, db: Session = Depends(get_db)):
    """Ventana de Reportes: tabla por estudiante (curso · prueba · ponderación · RUT · nombre ·
    puntaje · nota) con marca de qué estudiantes tienen detalle re-consultable."""
    try:
        return reportes_desarrollo_service.tabla_reportes(db, assessment_id)
    except Exception:
        logger.error(f"Error en reportes_desarrollo {assessment_id}: {traceback.format_exc()}")
        raise


@router.get("/assessments/{assessment_id}/desarrollo/reporte/{student_id}",
            dependencies=[Depends(req_profesor)])
def reporte_estudiante(assessment_id: UUID, student_id: UUID, db: Session = Depends(get_db)):
    """Detalle de la revisión de un estudiante: por pregunta, respuesta + revisión + criterios."""
    return reportes_desarrollo_service.detalle_estudiante(db, assessment_id, student_id)


@router.post("/assessments/{assessment_id}/desarrollo/reporte/{student_id}",
             dependencies=[Depends(req_profesor)])
def guardar_reporte_estudiante(assessment_id: UUID, student_id: UUID, payload: dict,
                               db: Session = Depends(get_db)):
    """Persiste (upsert) el detalle por pregunta de un estudiante (lo alimenta la corrección por
    lote). payload = {docente?, preguntas:[{item_id?, question_number|numero, respuesta, puntaje?,
    frac?, nivel?, revision?}]}."""
    preguntas = payload.get("preguntas") or []
    if not isinstance(preguntas, list):
        raise conflict("preguntas debe ser una lista.")
    return reportes_desarrollo_service.guardar_detalle(
        db, assessment_id, student_id, preguntas, docente=(payload.get("docente") or "docente"))


@router.post("/desarrollo/redactar", dependencies=[Depends(req_profesor)])
def redactar_desarrollo(payload: dict):
    """Limpia un texto dictado por el docente (enunciado/respuesta óptima): corrige transcripción
    y ortografía SIN cambiar el significado ni responder. payload = {texto, contexto?}."""
    return correccion_experta_service.redactar_texto(
        (payload.get("texto") or ""), (payload.get("contexto") or ""))


@router.get("/answer-keys/{ak_id}/rubrica/versiones", dependencies=[Depends(req_profesor)])
def listar_versiones(ak_id: UUID, db: Session = Depends(get_db)):
    """F4 - Historial de versiones de la rubrica (replicabilidad: cada corrida se clava a una)."""
    vers = (db.query(RubricaVersion)
            .filter(RubricaVersion.answer_key_id == str(ak_id))
            .order_by(RubricaVersion.version).all())
    return {"versiones": [{"version": v.version, "hash": v.hash, "parent_hash": v.parent_hash,
                           "estado": v.estado, "resumen": v.resumen, "autor": v.autor}
                          for v in vers]}


@router.post("/answer-keys/{ak_id}/rubrica/versiones/activar", dependencies=[Depends(req_profesor)])
def activar_version(ak_id: UUID, payload: dict, db: Session = Depends(get_db)):
    """
    F4 - Aplica los ajustes APROBADOS por el docente y activa una version NUEVA de la rubrica
    (la anterior queda archivada e inmutable). El docente aprueba la regla (G1 extendido); si
    un ajuste relaja la norma disciplinar, exige justificacion (se rechaza si falta).

    payload = {autor, aprobados:[{criterio, tipo, payload?, recurrencia?, requiere_override?,
               justificacion?, aprobado_por?}]}
    """
    try:
        autor = payload.get("autor", "docente")
        criterios = matriz_service.cargar_criterios_rubrica(db, ak_id)
        actual = (db.query(RubricaVersion)
                  .filter(RubricaVersion.answer_key_id == str(ak_id))
                  .order_by(RubricaVersion.version.desc()).first())
        nueva = aprendizaje_service.aplicar_ajustes(
            criterios, payload.get("aprobados", []),
            version_actual=(actual.version if actual else 1), autor=autor)

        db.query(RubricaVersion).filter(
            RubricaVersion.answer_key_id == str(ak_id),
            RubricaVersion.estado == "activa").update({"estado": "archivada"})
        db.add(RubricaVersion(answer_key_id=str(ak_id), version=nueva["version"],
                              hash=nueva["hash"], parent_hash=nueva["parent_hash"],
                              estado="activa", resumen=nueva["resumen"], autor=autor))
        for ch in nueva["changelog"]:
            db.add(AjusteCalibracion(
                rubrica_version_hash=nueva["hash"], criterio=ch["criterio"], tipo=ch["tipo"],
                direccion=ch.get("direccion") or "sube", descripcion=ch["tipo"],
                recurrencia=ch.get("recurrencia") or 1,
                requiere_override=bool(ch.get("requiere_override")),
                justificacion=ch.get("justificacion"), estado="aprobado",
                aprobado_por=ch.get("aprobado_por") or autor))
        db.commit()
        return {"version": nueva["version"], "hash": nueva["hash"], "estado": "activa",
                "n_cambios": nueva["n_cambios"], "resumen": nueva["resumen"],
                "gobernanza": nueva["gobernanza"]}
    except ValueError as e:                    # relaja la norma sin justificacion (G1)
        raise conflict(str(e))
    except Exception:
        logger.error(f"Error en activar_version {ak_id}: {traceback.format_exc()}")
        raise


@router.post("/answer-key-items/{item_id}/rubrica/importar", dependencies=[Depends(req_profesor)])
async def importar_rubrica(item_id: UUID, file: UploadFile = File(...), confirmar: bool = False,
                           aplicar_config: bool = True, db: Session = Depends(get_db)):
    """
    Importa una rubrica desde .xlsx a un item. Sin `confirmar` devuelve solo el PREVIEW (no
    escribe); con `confirmar=true` crea los criterios en el item. Si la planilla trae parametros
    de calificacion (exigencia, escala) y `aplicar_config`, tambien CONFIGURA la evaluacion
    (escala + exigencia) para que la nota salga identica a la planilla. Heuristico: el docente
    revisa y confirma (G1).
    """
    try:
        from app.services import rubrica_import_service
        from app.models.answer_key import AnswerKeyItem, AnswerKey, RubricCriterion
        from app.models.assessment import Assessment
        preview = rubrica_import_service.parse_rubrica_xlsx(await file.read())
        if not confirmar:
            return {"guardado": False, **preview}
        item = db.get(AnswerKeyItem, item_id)
        if not item:
            raise not_found("Item de pauta no encontrado.")
        # Reemplazo total: borra los criterios previos (y sus anclas por cascada) para que
        # reimportar no duplique; el docente parte de la planilla nueva.
        for _old in db.query(RubricCriterion).filter(
                RubricCriterion.answer_key_item_id == item_id).all():
            db.delete(_old)
        db.flush()
        for i, c in enumerate(preview["criterios"]):
            db.add(RubricCriterion(
                answer_key_item_id=item_id, name=c["name"][:255], weight=c["weight"],
                order=i, niveles_json=c["niveles"], ambito=c["ambito"],
                seccion=(c["seccion"][:120] if c["seccion"] else None)))
        # Configurar la evaluacion desde la planilla (escala + exigencia).
        config_aplicada = None
        cfg = preview.get("config") or {}
        if aplicar_config and cfg:
            ak = db.get(AnswerKey, item.answer_key_id)
            a = db.get(Assessment, ak.assessment_id) if ak else None
            if a:
                aplicada = {}
                if cfg.get("escala"):
                    a.grading_scale = cfg["escala"]; aplicada["escala"] = cfg["escala"]
                if cfg.get("exigencia") is not None:
                    a.passing_threshold = float(cfg["exigencia"]); aplicada["exigencia"] = cfg["exigencia"]
                if aplicada:
                    db.add(a); config_aplicada = aplicada
        db.commit()
        return {"guardado": True, "criterios_creados": len(preview["criterios"]),
                "config_aplicada": config_aplicada, **preview}
    except Exception:
        logger.error(f"Error en importar_rubrica {item_id}: {traceback.format_exc()}")
        raise


_NIVELES = ("logrado", "parcial", "no_logrado")


def _serializa_criterio(c) -> dict:
    """Criterio en forma editable para el editor del docente."""
    anclas = sorted(c.anclas, key=lambda a: (a.nivel, a.order))
    return {
        "id": str(c.id), "name": c.name, "descriptor": c.descriptor,
        "weight": float(c.weight), "order": c.order,
        "nivel_exigencia": c.nivel_exigencia, "umbral_confianza": float(c.umbral_confianza),
        "penaliza_forma": bool(c.penaliza_forma), "sinonimos": c.sinonimos_json or [],
        "fuera_de_alcance": c.fuera_de_alcance, "seccion": c.seccion, "ambito": c.ambito,
        "niveles": c.niveles_json or None,   # escala propia N-niveles (Excelente=3/…); None = 3 niveles por defecto
        "anclas": [{"texto": a.texto, "nivel": a.nivel} for a in anclas],
    }


@router.get("/answer-key-items/{item_id}/rubrica", dependencies=[Depends(req_profesor)])
def obtener_rubrica(item_id: UUID, db: Session = Depends(get_db)):
    """Devuelve la rúbrica editable de un ítem de desarrollo (criterios + anclas por nivel)."""
    from app.models.answer_key import AnswerKeyItem
    item = db.get(AnswerKeyItem, item_id)
    if not item:
        raise not_found("Ítem de pauta no encontrado.")
    crit = sorted(item.rubric_criteria, key=lambda x: x.order)
    return {"item_id": str(item_id), "criterios": [_serializa_criterio(c) for c in crit]}


@router.put("/answer-key-items/{item_id}/rubrica", dependencies=[Depends(req_profesor)])
def guardar_rubrica(item_id: UUID, payload: dict, db: Session = Depends(get_db)):
    """
    Reemplaza la rúbrica del ítem con la que define el docente (autoría manual, G1).
    payload = {criterios:[{name, weight, nivel_exigencia, umbral_confianza, seccion, ambito,
               sinonimos:[...], descriptor, fuera_de_alcance, anclas:[{texto, nivel}]}]}
    Valida: al menos 1 criterio con nombre; pesos > 0; niveles de ancla en {logrado,parcial,no_logrado}.
    """
    from app.models.answer_key import AnswerKeyItem, RubricCriterion, RubricAncla
    item = db.get(AnswerKeyItem, item_id)
    if not item:
        raise not_found("Ítem de pauta no encontrado.")
    criterios = payload.get("criterios", [])
    limpios = [c for c in criterios if (c.get("name") or "").strip()]
    if not limpios:
        raise conflict("La rúbrica necesita al menos un criterio con nombre.")
    for c in limpios:
        for a in c.get("anclas", []):
            if a.get("nivel") not in _NIVELES:
                raise conflict(f"Nivel de ancla inválido: {a.get('nivel')} (usa logrado/parcial/no_logrado).")
    # Reemplazo total (los criterios y sus anclas viejos se borran en cascada).
    for old in list(item.rubric_criteria):
        db.delete(old)
    db.flush()
    for i, c in enumerate(limpios):
        try:
            peso = float(c.get("weight", 1.0))
        except (TypeError, ValueError):
            peso = 1.0
        if peso <= 0:
            raise conflict(f"El peso del criterio '{c.get('name')}' debe ser > 0.")
        crit = RubricCriterion(
            answer_key_item_id=item_id, name=str(c["name"])[:255],
            descriptor=(c.get("descriptor") or None), weight=peso, order=i,
            nivel_exigencia=c.get("nivel_exigencia") or "tolerante",
            penaliza_forma=bool(c.get("penaliza_forma", False)),
            sinonimos_json=(c.get("sinonimos") or []),
            umbral_confianza=float(c.get("umbral_confianza", 0.7)),
            fuera_de_alcance=(c.get("fuera_de_alcance") or None),
            seccion=((c.get("seccion") or None) and str(c["seccion"])[:120]),
            ambito=c.get("ambito") or "individual",
            niveles_json=(c.get("niveles") or None))   # preserva la escala N-niveles importada
        db.add(crit)
        db.flush()
        for j, a in enumerate(c.get("anclas", [])):
            if (a.get("texto") or "").strip():
                db.add(RubricAncla(rubric_criterion_id=crit.id, texto=str(a["texto"]),
                                   nivel=a["nivel"], order=j))
    db.commit()
    db.refresh(item)
    crit = sorted(item.rubric_criteria, key=lambda x: x.order)
    return {"guardado": True, "n_criterios": len(crit),
            "criterios": [_serializa_criterio(c) for c in crit]}


# ───────────────────────────── Instrumento de desarrollo MULTI-PREGUNTA (peso por puntaje)
def _pregunta_dict(it, total_peso) -> dict:
    return {"id": str(it.id), "numero": it.question_number, "enunciado": it.enunciado or "",
            "weight": float(it.weight or 1),
            "respuesta_optima": (it.respuesta_optima if it.respuesta_optima is not None else (it.correct_answer or "")),
            "nivel_rigor": getattr(it, "nivel_rigor", None) or "estricto",
            "area_conocimiento": getattr(it, "area_conocimiento", None) or "general",
            "fuente_estandar": getattr(it, "fuente_estandar", None) or "",
            "pct": round(float(it.weight or 0) / total_peso * 100, 1) if total_peso else 0.0,
            "n_criterios": len(it.rubric_criteria)}


@router.get("/answer-keys/{ak_id}/desarrollo", dependencies=[Depends(req_profesor)])
def listar_preguntas_desarrollo(ak_id: UUID, db: Session = Depends(get_db)):
    """Preguntas de desarrollo (open_response) de la pauta, con enunciado, peso (puntaje) y % del total."""
    from app.models.answer_key import AnswerKey, QUESTION_TYPE_OPEN_RESPONSE
    ak = db.get(AnswerKey, ak_id)
    if not ak:
        raise not_found("Pauta no encontrada.")
    items = [it for it in sorted(ak.items, key=lambda x: x.question_number)
             if it.question_type == QUESTION_TYPE_OPEN_RESPONSE]
    total = sum(float(it.weight or 0) for it in items) or 1.0
    return {"answer_key_id": str(ak.id), "n_preguntas": len(items), "puntaje_total": round(total, 1),
            "preguntas": [_pregunta_dict(it, total) for it in items]}


@router.post("/answer-keys/{ak_id}/desarrollo", dependencies=[Depends(req_profesor)])
def crear_pregunta_desarrollo(ak_id: UUID, payload: dict, db: Session = Depends(get_db)):
    """Crea una pregunta de desarrollo. payload = {enunciado, weight(puntaje)}."""
    from app.models.answer_key import AnswerKey, AnswerKeyItem, QUESTION_TYPE_OPEN_RESPONSE
    ak = db.get(AnswerKey, ak_id)
    if not ak:
        raise not_found("Pauta no encontrada.")
    try:
        w = float(payload.get("weight", 1) or 1)
    except (TypeError, ValueError):
        w = 1.0
    nums = [it.question_number for it in ak.items] or [0]
    it = AnswerKeyItem(answer_key_id=ak.id, question_number=max(nums) + 1, version="A",
                       correct_answer="",
                       respuesta_optima=((payload.get("respuesta_optima") or "").strip() or None),
                       weight=max(0.1, w), is_annulled=False,
                       question_type=QUESTION_TYPE_OPEN_RESPONSE,
                       enunciado=((payload.get("enunciado") or "").strip() or None))
    db.add(it)
    ak.is_valid = True
    db.commit(); db.refresh(it)
    return _pregunta_dict(it, float(it.weight or 1))


@router.patch("/answer-key-items/{item_id}/desarrollo", dependencies=[Depends(req_profesor)])
def editar_pregunta_desarrollo(item_id: UUID, payload: dict, db: Session = Depends(get_db)):
    """Edita enunciado, peso o número (reordenar) de una pregunta de desarrollo."""
    from app.models.answer_key import AnswerKeyItem
    it = db.get(AnswerKeyItem, item_id)
    if not it:
        raise not_found("Pregunta no encontrada.")
    if "enunciado" in payload:
        it.enunciado = (payload.get("enunciado") or "").strip() or None
    if "respuesta_optima" in payload:
        it.respuesta_optima = (payload.get("respuesta_optima") or "").strip() or None
    if "nivel_rigor" in payload:
        from app.models.answer_key import NIVELES_RIGOR
        v = str(payload.get("nivel_rigor") or "").strip()
        if v in NIVELES_RIGOR:
            it.nivel_rigor = v
    if "area_conocimiento" in payload:
        it.area_conocimiento = (str(payload.get("area_conocimiento") or "").strip() or None)
    if "fuente_estandar" in payload:
        it.fuente_estandar = (str(payload.get("fuente_estandar") or "").strip() or None)
    if "weight" in payload:
        try:
            w = float(payload["weight"])
            if w > 0:
                it.weight = w
        except (TypeError, ValueError):
            pass
    if "question_number" in payload:
        try:
            it.question_number = int(payload["question_number"])
        except (TypeError, ValueError):
            pass
    db.commit()
    return _pregunta_dict(it, float(it.weight or 1))


@router.delete("/answer-key-items/{item_id}/desarrollo", dependencies=[Depends(req_profesor)])
def borrar_pregunta_desarrollo(item_id: UUID, db: Session = Depends(get_db)):
    """Elimina una pregunta de desarrollo (y su rúbrica en cascada)."""
    from app.models.answer_key import AnswerKeyItem
    it = db.get(AnswerKeyItem, item_id)
    if not it:
        raise not_found("Pregunta no encontrada.")
    db.delete(it)
    db.commit()
    return {"borrado": True}


@router.post("/answer-keys/{ak_id}/desarrollo/importar", dependencies=[Depends(req_profesor)])
async def importar_prueba_desarrollo(ak_id: UUID, file: UploadFile = File(...),
                                     confirmar: bool = False, db: Session = Depends(get_db)):
    """
    Importa la PRUEBA de desarrollo desde .xlsx: una fila por pregunta con columnas
    'enunciado' (o 'pregunta') y 'peso' (o 'puntaje'). Sin `confirmar` devuelve preview;
    con confirmar=true crea las preguntas (G1: el docente revisa antes).
    """
    from app.models.answer_key import AnswerKey, AnswerKeyItem, QUESTION_TYPE_OPEN_RESPONSE
    from openpyxl import load_workbook
    import io as _io
    ak = db.get(AnswerKey, ak_id)
    if not ak:
        raise not_found("Pauta no encontrada.")
    wb = load_workbook(_io.BytesIO(await file.read()), read_only=True, data_only=True)
    ws = wb.active
    filas = list(ws.iter_rows(values_only=True))
    if not filas:
        return {"guardado": False, "preguntas": [], "nota": "Planilla vacía."}
    cab = [str(c).strip().lower() if c is not None else "" for c in filas[0]]
    def col(*names):
        for n in names:
            if n in cab:
                return cab.index(n)
        return None
    ci_en = col("enunciado", "pregunta", "enunciado de la pregunta")
    ci_pe = col("peso", "puntaje", "puntos", "ponderacion", "ponderación")
    ci_ro = col("respuesta_optima", "respuesta optima", "respuesta óptima", "respuesta de referencia", "referencia")
    if ci_en is None:
        return {"guardado": False, "preguntas": [], "nota": "Falta la columna 'enunciado' (o 'pregunta')."}
    prev = []
    for row in filas[1:]:
        en = row[ci_en] if ci_en < len(row) else None
        if en is None or not str(en).strip():
            continue
        try:
            pe = float(row[ci_pe]) if (ci_pe is not None and ci_pe < len(row) and row[ci_pe] is not None) else 1.0
        except (TypeError, ValueError):
            pe = 1.0
        ro = ""
        if ci_ro is not None and ci_ro < len(row) and row[ci_ro] is not None:
            ro = str(row[ci_ro]).strip()
        prev.append({"enunciado": str(en).strip(), "weight": max(0.1, pe), "respuesta_optima": ro})
    if not confirmar:
        return {"guardado": False, "n": len(prev), "preguntas": prev,
                "nota": "Vista previa. Confirma para crear las preguntas."}
    nums = [it.question_number for it in ak.items] or [0]
    n0 = max(nums)
    creadas = []
    for i, p in enumerate(prev, start=1):
        it = AnswerKeyItem(answer_key_id=ak.id, question_number=n0 + i, version="A",
                           correct_answer="", respuesta_optima=(p.get("respuesta_optima") or None),
                           weight=p["weight"], is_annulled=False,
                           question_type=QUESTION_TYPE_OPEN_RESPONSE, enunciado=p["enunciado"])
        db.add(it); creadas.append(p)
    ak.is_valid = True
    db.commit()
    return {"guardado": True, "creadas": len(creadas), "preguntas": creadas}


@router.get("/assessments/{assessment_id}/rubrica/mfrm", dependencies=[Depends(req_investigador)])
def rubrica_mfrm(assessment_id: UUID, db: Session = Depends(get_db)):
    """I6 - Severidad del corrector IA vs docente (MFRM) sobre las validaciones persistidas."""
    try:
        regs = matriz_service.cargar_registros_validacion(db, assessment_id)
        return mfrm_service.reporte_severidad_ia(regs)
    except Exception:
        logger.error(f"Error en rubrica_mfrm {assessment_id}: {traceback.format_exc()}")
        raise


@router.get("/assessments/{assessment_id}/rubrica/psicometria", dependencies=[Depends(req_investigador)])
def rubrica_psicometria(assessment_id: UUID, db: Session = Depends(get_db)):
    """R - Psicometria de la rubrica (estadigrafos por criterio, fiabilidad, categorias, G-theory)."""
    try:
        regs = matriz_service.cargar_registros_validacion(db, assessment_id)
        return rubrica_psicometria_service.analizar_rubrica(regs)
    except Exception:
        logger.error(f"Error en rubrica_psicometria {assessment_id}: {traceback.format_exc()}")
        raise


@router.get("/assessments/{assessment_id}/rubrica/aprendizaje", dependencies=[Depends(req_profesor)])
def rubrica_aprendizaje(assessment_id: UUID, db: Session = Depends(get_db)):
    """F4 - Propuestas de ajuste aprendidas del docente (solo propone; el docente aprueba, G1)."""
    try:
        regs = matriz_service.cargar_registros_validacion(db, assessment_id)
        return aprendizaje_service.ciclo_aprendizaje(regs)
    except Exception:
        logger.error(f"Error en rubrica_aprendizaje {assessment_id}: {traceback.format_exc()}")
        raise
